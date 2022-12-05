from PyPDF2 import PdfReader, PdfFileWriter, PdfFileReader
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter #get the column letter
from reportlab.pdfgen.canvas import Canvas
import io
import shutil  #Mover arquivos para diretórios
import os  #Functions to access, modify, and perform OS-related tasks such as access and modifying directories.
from PIL import Image, ImageFont, ImageDraw
from PyPDF4 import PdfFileWriter, PdfFileReader 

def coletarInfosPDF(nome_pdf):
    file = open(nome_pdf, 'rb')
    read = PdfReader(file)
    first_page = read.getPage(0).extractText()

    #Base para verificar se as informações estão corretas:
    ender_CAEP = 'R JOSE GERALDO CEREBINO CHRISTOFARO, 245 - PARQUE RURAL FAZENDA SANTA CANDIDA - CAMPINAS - SP, CEP.: 13087-567'
    dados_equipamento = {'Nº CONTROLE' : '',
                         'Nº SÉRIE' : '', 
                         'MARCA' : '',
                         'MODELO' : '',
                         'DATA DA CALIBRAÇÃO' : ''}
  
    def buscar_dado(dado, line):
        if dado in line:
            aux = ""
            caracter = 0  #Só para declarar, linha 26 que importa
            now = False
            for c in line:
                aux += c
                if c == ":" and dado in aux:    #Próximo valor de 'c' é o inicio da informação
                    caracter = -1               #Ignorar espaço em branco entre letras e numeros na TAG
                    aux = ""
                    now = True          
                if now == True and caracter > 2:
                    if c == " " or c == "\n":   #Fim da informação
                        break
                caracter += 1        
            dados_equipamento[dado] = aux.upper() #Todas as letras serão upper para facilitar a comparação dos dados
    linha = ""
    ender_CAEP_achou = False
    l = 1                       #Contador de linhas
    for caracter in first_page: #Percorre todos caracteres do pdf
        if caracter != '\n':
            linha += caracter   #Adiciona caracteres em 'linha' até encontrar o 'enter', vulgo '\n'
            
        else:                   #Fim da linha, hora de procurar as informações nela:
            if ender_CAEP in linha:
                ender_CAEP_achou = True
                print("Endereço CAEP está correto! Encontrado na linha " + str(l)) 
                     
            buscar_dado('Nº CONTROLE', linha)
            buscar_dado('MARCA', linha)
            buscar_dado('Nº SÉRIE', linha)
            buscar_dado('MODELO', linha)
            buscar_dado('DATA DA CALIBRAÇÃO', linha)
            linha = ""
            l += 1     
    file.close()   
             
    if ender_CAEP_achou == False:
        print("Endereço CAEP não encontrado!") 
                   
    print(dados_equipamento)  #REMOVER ESSA LINHA!!
    return dados_equipamento
#------------------------------------------------------------------------------------------------------------------------------#  
def buscarPlanilha(p, dadosPDF):
    workbook = load_workbook(p)
    #ws = workbook['Planilha1']   #Acesse diretamente o Sheet 1
    sheet = workbook.active       #Ou use a aba ativa quando o arquivo foi carregado
    n_linhas = sheet.max_row
    n_colunas = sheet.max_column
    
    alfabeto = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    dados = ['NOME', 'TAG', 'NÚMERO DE SÉRIE', 'MARCA / MODELO', 'DATA', 'SITUAÇÃO']
    
    #Descobrir quais as colunas que interessam (Obs: o titulo pode estar na 1º ou 2º linha):
    colunas = []
    
    for i in range(n_colunas):
        if sheet[alfabeto[i] + '1'].value != None:    
            for dado in dados:
                if dado in sheet[alfabeto[i] + '1'].value:
                    colunas.append(alfabeto[i] + '1')
                
        if sheet[alfabeto[i] + '2'].value != None:
            for dado in dados:         
                if dado in sheet[alfabeto[i] + '2'].value:
                    colunas.append(alfabeto[i] + '2')
                    
    #Buscar o item do PDF na planilha, linha por linha, usando as colunas que interessam:
    if colunas[1] == '1':
        first_item = 1
    else:
        first_item = 2
        
    TAG = False; SERIE = False; MARCA = False; DATA = False; FLAG = False
    print("\n--> Buscando item na Planilha: <--") #REMOVER ESSA LINHA!!
    for linha in range(first_item, n_linhas):
        d = 0  #Indice para representar a lista 'dados'  
          
        for coluna in colunas:
            index = str(coluna[0]) + str(linha + 1)  #coluna[0] representa a letra das colunas que importam
            
            #Formatando datas:
            if sheet[index].value != None and d == 4 and type(sheet[index].value) != str:
                item = str(format(sheet[index].value, '%d/%m/%y'))
            else:
                item = str(sheet[index].value).upper() 
                
            #Validando itens: 
            if d == 1: #TAG
                if " " in item and "-" not in item:  #Garantindo que a TAG segue o padrão XX-111
                    item = item.replace(' ','-')
                if item in dadosPDF['Nº CONTROLE']:
                    print('TAG ' + item + ' ENCONTRADA!')
                    TAG = True     
            if d == 2 and TAG == True: #Nº SÉRIE
                if item in dadosPDF['Nº SÉRIE']:
                    print('Nº SÉRIE ' + item + ' ENCONTRADO!')
                    SERIE = True
                elif item == "NA" and dadosPDF['Nº SÉRIE'] == "NÃO ":
                    print("O item NÃO possui Nº SÉRIE!")
                    SERIE = True                        
            if d == 3 and SERIE == True: #MARCA/MODELO
                if item in dadosPDF['MARCA']:
                    print("MARCA/MODELO " + item + " ENCONTRADO!")
                    MARCA = True
                elif dadosPDF['MARCA'] == "NA":
                    print("O item NÃO possui MARCA/MODELO")
                    MARCA = True
            if d == 4 and MARCA == True: #DATA
                if item in dadosPDF['DATA DA CALIBRAÇÃO']:
                    print("DATA DA CALIBRAÇÃO " + item + " ENCONTRADA!")
                    DATA = True
            if d == 5 and DATA == True:
                print("SITUAÇÃO: " + item)
                FLAG = True
                break
            d += 1
        #Obs:'item in' foi usado ao invés de 'item ==' para driblar, na condição(if), possíveis espaços em branco no final de strings     
        
        if FLAG == True: #Os dados foram encontrados na Planilha
            return True

    return False #Dados não bateram        
#------------------------------------------------------------------------------------------------------------------------------#  
def assinarPDF(nome, pdf, data):
    ass = "Conferido por: " + nome + " | " + data
    caminho = os.path.expanduser("~") + '\Downloads' + '\\' + data

    page_to_merge = 1 #Refers to the First page of PDF 
    input_pdf = PdfFileReader(open(caminho  + '\Certificados\\' + pdf, "rb"))
    page_count = input_pdf.getNumPages()
    inputpdf_page_to_be_merged = input_pdf.getPage(page_to_merge)

    packet = io.BytesIO()
    c = Canvas(packet,pagesize=(inputpdf_page_to_be_merged.mediaBox.getWidth(),inputpdf_page_to_be_merged.mediaBox.getHeight()))

    x = int(inputpdf_page_to_be_merged.mediaBox.getWidth())*0.07
    y = int(inputpdf_page_to_be_merged.mediaBox.getHeight())*0.28

    c.drawString(x, y, ass)
    c.save()
    packet.seek(0)

    overlay_pdf = PdfFileReader(packet)
    overlay = overlay_pdf.getPage(0)

    output = PdfFileWriter()

    for PAGE in range(page_count):
        if PAGE == page_to_merge:
            inputpdf_page_to_be_merged.mergeRotatedTranslatedPage(overlay, 
                    inputpdf_page_to_be_merged.get('/Rotate') or 0, 
                    int(overlay.mediaBox.getWidth())/2, int(overlay.mediaBox.getHeight())/2)
            output.addPage(inputpdf_page_to_be_merged)
        
        else:
            Page_in_pdf = input_pdf.getPage(PAGE)
            output.addPage(Page_in_pdf)
    
    outputStream = open(caminho + "\Assinados\\" + pdf, "wb")
    output.write(outputStream)
    outputStream.close()
    addWatermark(caminho + "\Assinados\\" + pdf, 'watermark.pdf')
#------------------------------------------------------------------------------------------------------------------------------#
def addWatermark(pdf, watermark): 
    watermark_instance = PdfFileReader(watermark) 
    watermark_page = watermark_instance.getPage(0) 
    pdf_reader = PdfFileReader(pdf) 
    pdf_writer = PdfFileWriter()

    for page in range(pdf_reader.getNumPages()): 
        page = pdf_reader.getPage(page)   
        page.mergePage(watermark_page) 
        pdf_writer.addPage(page) 

    with open(pdf, 'wb') as out: #output_pdf
        pdf_writer.write(out)

#addWatermark('LV01061-32983-22-R0.pdf', 'watermark.pdf')