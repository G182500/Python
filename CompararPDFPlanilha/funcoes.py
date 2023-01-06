from PyPDF2 import PdfReader, PdfFileWriter, PdfFileReader, PdfWriter
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter #get the column letter
from reportlab.pdfgen.canvas import Canvas
import io
import os  #Functions to access, modify, and perform OS-related tasks such as access and modifying directories.
from bd import salvarBanco
from datetime import datetime
import configparser

#CONFIG
config = configparser.ConfigParser()
config.read('python_config.ini')

#---------------------------------------------------------------------------------------------------------------------------
def coletarInfosPDF(nome_pdf):
    file = open(nome_pdf, 'rb')
    read = PdfReader(file)
    first_page = read.pages[0].extract_text()
    
    #Base para verificar se as informações estão corretas:
    dados_equipamento = {'Nº CONTROLE' : '',
                         'Nº SÉRIE' : '', 
                         'MARCA' : '',
                         'MODELO' : '',
                         'DATA DA CALIBRAÇÃO' : ''}
  
    def buscar_dado(dado, line):
        if dado in line:
            aux = ""; now = False; caracter = 0
            
            for c in line:
                aux += c
                
                if c == ":" and dado in aux:             #Próximo valor de 'c' é o inicio da informação
                    aux = ""; now = True; caracter = -1  #Ignorar espaço em branco entre letras e numeros na TAG            
                
                if now == True and caracter > 2:
                    if c == " " or c == "\n":            #Fim da informação
                        break
                
                caracter += 1        
            dados_equipamento[dado] = aux.upper() #Todas as letras serão upper para facilitar a comparação dos dados
    
    linha = ""; l = 1           #Contador de linhas
    
    for caracter in first_page: #Percorre todos caracteres do pdf
        if caracter != '\n':
            linha += caracter   #Adiciona caracteres em 'linha' até encontrar o 'enter', vulgo '\n' 
        
        else:                   #Fim da linha, hora de procurar as informações nela:       
            buscar_dado('Nº CONTROLE', linha)
            buscar_dado('MARCA', linha)
            buscar_dado('Nº SÉRIE', linha)
            buscar_dado('MODELO', linha)
            buscar_dado('DATA DA CALIBRAÇÃO', linha)
            linha = ""
            l += 1     
    
    file.close()              
    return dados_equipamento

#---------------------------------------------------------------------------------------------------------------------------
def buscarPlanilha(p, dadosPDF, nome_pdf):
    workbook = load_workbook(p)
    
    #ws = workbook['Planilha1']   #Acesse diretamente o Sheet 1
    sheet = workbook.active       #Ou use a aba ativa quando o arquivo foi carregado
    
    n_linhas = sheet.max_row; n_colunas = sheet.max_column
    alfabeto = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    dados = ['NOME', 'TAG', 'NÚMERO DE SÉRIE', 'MARCA / MODELO', 'DATA', 'SITUAÇÃO']
    
    #Descobrir quais as colunas que interessam (Obs: o titulo pode estar na 1º ou 2º linha):
    colunas = []
    for i in range(n_colunas):
        if sheet[alfabeto[i] + '1'].value != None:    
            
            for dado in dados:
                if dado in sheet[alfabeto[i] + '1'].value:
                    colunas.append(alfabeto[i] + '1')         
        
        if sheet[alfabeto[i] + '2'].value != None:
            
            for dado in dados:         
                if dado in sheet[alfabeto[i] + '2'].value:
                    colunas.append(alfabeto[i] + '2')
                    
    #Buscar o item do PDF na planilha, linha por linha, usando as colunas que interessam:
    if colunas[1] == '1':
        first_item = 1
    else:
        first_item = 2
   
    TAG = "Não encontrado"; SERIE = "Não encontrado"; MARCA = "Não encontrado"; DATA = "Não encontrado"; SITUACAO = "Não encontrado"; FLAG = False
    #print("\n--> Buscando item na Planilha: <--")
    
    for linha in range(first_item, n_linhas):
        d = 0  #Indice para representar a lista 'dados'  
        
        for coluna in colunas:
            index = str(coluna[0]) + str(linha + 1)  #coluna[0] representa a letra das colunas que importam
            
            #Formatando datas:
            if sheet[index].value != None and d == 4 and type(sheet[index].value) != str:
                item = str(format(sheet[index].value, '%d/%m/%y'))
            else:
                item = str(sheet[index].value).upper() 
            
            #Validando itens: 
            if d == 1: #TAG
                if " " in item and "-" not in item:  #Garantindo que a TAG segue o padrão XX-111
                    item = item.replace(' ','-')
                if item in dadosPDF['Nº CONTROLE']:
                    print('TAG ' + item + ' ENCONTRADA!')
                    TAG = item     
           
            if d == 2 and TAG != "Não encontrado": #Nº SÉRIE
                if item in dadosPDF['Nº SÉRIE']:
                    print('Nº SÉRIE ' + item + ' ENCONTRADO!')
                    SERIE = item
                elif item == "NA" and dadosPDF['Nº SÉRIE'] == "NÃO ":
                    print("O item NÃO possui Nº SÉRIE!")
                    SERIE = item                        
            
            if d == 3 and SERIE != "Não encontrado": #MARCA/MODELO
                if item in dadosPDF['MARCA']:
                    print("MARCA/MODELO " + item + " ENCONTRADO!")
                    MARCA = item
                elif dadosPDF['MARCA'] == "NA":
                    print("O item NÃO possui MARCA/MODELO")
                    MARCA = item
            
            if d == 4 and MARCA != "Não encontrado": #DATA
                if item in dadosPDF['DATA DA CALIBRAÇÃO']:
                    print("DATA DA CALIBRAÇÃO " + item + " ENCONTRADA!")
                    DATA = item
           
            if d == 5 and DATA != "Não encontrado":
                SITUACAO = item
                FLAG = True
                break
            
            d += 1
        #Obs:'item in' foi usado ao invés de 'item ==' para driblar, na condição(if), possíveis espaços em branco no final de strings     
        
        infos = "TAG = " + TAG + "; SERIE = " + SERIE + "; MARCA = " + MARCA + "; DATA = " + DATA + "; SITUAÇÃO = " + SITUACAO
        
        #Salvar no BD:
        if FLAG == True: #Os dados foram encontrados na Planilha
            salvarBanco(nome_pdf, infos, status = 'PRESENTE NA PLANILHA')
            return [True, TAG]

    salvarBanco(nome_pdf, infos, status = 'PENDENTE')
    return [False, None] #Dados não bateram        

#---------------------------------------------------------------------------------------------------------------------------
def addWatermark(pdf, watermark): 
    watermark_instance = PdfReader(watermark) 
    watermark_page = watermark_instance.pages[0]
    pdf_reader = PdfReader(pdf) 
    pdf_writer = PdfWriter()

    for page in range(len(pdf_reader.pages)): 
        page = pdf_reader.pages[page]   
        page.merge_page(watermark_page) 
        pdf_writer.add_page(page) 

    with open(pdf, 'wb') as out: #output_pdf
        pdf_writer.write(out)

    out.close()
    
#---------------------------------------------------------------------------------------------------------------------------
def assinarPDF(nome, pdf, data):
    sharepoint = os.path.expanduser("~") + config['CAMINHOS']['assinados'] 

    page_to_merge = 1 #Refers to the First page of PDF 
    input_pdf = PdfReader(open(sharepoint + '\\' + pdf, "rb"))
    
    page_count = len(input_pdf.pages)
    inputpdf_page_to_be_merged = input_pdf.pages[page_to_merge]
    packet = io.BytesIO()
    
    mediabox_width = inputpdf_page_to_be_merged.mediabox.width
    mediabox_height = inputpdf_page_to_be_merged.mediabox.height

    c = Canvas(packet,pagesize=(mediabox_width,mediabox_height))
    x = int(mediabox_width)*0.07
    y = int(mediabox_height)*0.28

    c.drawString(x, y, "Conferido por: " + nome + " | " + data)
    c.save()
    packet.seek(0)

    overlay_pdf = PdfReader(packet)
    overlay = overlay_pdf.pages[0]
    output =  PdfWriter()

    for PAGE in range(page_count):
        if PAGE == page_to_merge:
            inputpdf_page_to_be_merged.merge_page(overlay)
            output.add_page(inputpdf_page_to_be_merged)
        else:
            Page_in_pdf = input_pdf.pages[PAGE]
            output.add_page(Page_in_pdf)
    
    outputStream = open(sharepoint + "\\" + pdf, "wb")
    output.write(outputStream)
    outputStream.close()
    
    addWatermark(sharepoint + "\\" + pdf, 'watermark.pdf')